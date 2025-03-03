from flask_api import db, ma, app
from flask import jsonify, Blueprint, render_template, flash, session, make_response, redirect, url_for
from flask_restful import request
from flask_api.models.client import Client
from flask_api.models.user_type import UserType
from flask_api.utils.get_id_from_db_object_for_relation import get_id, get_name_from_type
from flask_api.utils.generate_client_code import generate_code
from flask_api.auth_forms.client_form import RegForm, LoginForm
from flask_api.utils.generate_flask_jwt_auth_token import generate_token, update_token_expiry, verify_token
from flask_login import login_user, logout_user, current_user, login_required
from io import BytesIO
from PIL import Image
import re
import base64

main = Blueprint('client', __name__)


class ClientSchema(ma.Schema):
    class Meta:
        fields = ('id', 'first_name', 'last_name', 'phone',
                  'email', 'photo', 'client_code', 'type',
                  'registered', 'auth_token')


client_schema = ClientSchema()
client_schema = ClientSchema(many=True)


@main.route('/get_clients', methods=['GET'])
def get_clients():  # done
    clients = Client.query.all()
    output = []
    for client in clients:
        client_data = {}
        client_data['id'] = client.id
        client_data['first_name'] = client.first_name
        client_data['last_name'] = client.last_name
        client_data['phone'] = client.phone
        client_data['email'] = client.email
        client_data['photo'] = client.photo
        client_data['type_name'] = get_name_from_type(UserType, client.type_id)
        client_data['registered'] = client.registered

        output.append(client_data)

    return jsonify({'clients': output})


@main.route('/client/<id>', methods=['GET'])
def get_client(id):  # done
    client = Client.query.get(id)

    if not client:
        return {"message": "client with that id not found"}, 404

    output = []

    client_data = {}

    client_data['id'] = client.id
    client_data['first_name'] = client.first_name
    client_data['last_name'] = client.last_name
    client_data['phone'] = client.phone
    client_data['email'] = client.email
    client_data['photo'] = client.photo
    client_data['type_name'] = get_name_from_type(UserType, client.type_id)
    client_data['registered'] = client.registered

    output.append(client_data)

    return jsonify({'client': output})


@main.route('/add', methods=['GET', 'POST'])
def add_client():  # done
    form = RegForm(request.form)
    if request.method == "POST" and form.validate():
        phone = form.phone.data
        if phone.startswith("+38"):
            phone = phone[3:]  # Видаляємо перші 3 символи (+38)

        _phone = Client.query.filter_by(phone=phone).first()
        if _phone:
            flash("Find the same phone. Choose some any.", "error")
            return render_template('add_client.html', form=form)

        email = form.email.data
        if not is_valid_email(email):
            flash("Invalid email format.", "error")
            return render_template('add_client.html', form=form)

        _email = Client.query.filter_by(email=email).first()
        if _email:
            flash("Find the same email. Choose some any.", "error")
            return render_template('add_client.html', form=form)

        new_client = Client(
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            phone=phone,
            email=email,
            photo=None,
            client_code=generate_code(),
            type_id=get_id(UserType, form.name.data),
            auth_token=''
        )

        db.session.add(new_client)
        db.session.commit()

        flash("Client successfully added. Please login.", "success")
        return redirect(url_for('client.login'))

    return render_template('add_client.html', form=form)


@main.route('/update_client', methods=['GET', 'POST'])
def update_client():
    if request.method == 'POST':
        # field for identification client
        auth_token = request.form.get('auth_token')
        # fields that will be changed
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        phone = request.form.get('phone')

        client = Client.query.filter_by(auth_token=auth_token).first()

        if client:
            client.first_name = first_name
            client.last_name = last_name
            client.phone = phone

            db.session.commit()

            response_data = {
                "message": "profile change successful",
                "first_name": first_name,
                "last_name": last_name,
                "phone": phone
            }

            return jsonify(response_data), 200
        
    return {"message": "client does not found"}, 404



from flask_api.routes.login_data import add_code
from flask_api.models.login_data import LoginData
from sqlalchemy import desc


@main.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json()
        email = data["email"]

        if not email:
            return {"error": "Email is missing"}, 400

        delete_expired_records()
        # Перевірка і логіка для перевірки email
        client = Client.query.filter_by(email=email).first()
        if client:
            # Видаляються записи яким більше 1 години
            delete_expired_records()

            existing_code = LoginData.query.filter_by(client_id=client.id).order_by(
                LoginData.creation_time.desc()).first()

            if existing_code and not is_code_expired(existing_code.due_time):
                # Якщо користувач вже має активний код, надсилаємо його замість генерації нового коду
                send_esputnik_email(email, existing_code.code)
            else:
                add_code(email)
                new_code = LoginData.query.order_by(LoginData.code.desc()).first()
                send_esputnik_email(email, new_code.code)

            # generate auth token
            auth_token = generate_token(client.id)
            # update client auth token in the database
            client.auth_token = auth_token
            db.session.commit()
            # login the user
            login_user(client)
            # Перенаправлення на сторінку введення коду

            response_data = {
                "message": "email is correct",
                "auth_token": auth_token
            }

            return jsonify(response_data), 200
        else:
            return {"error": "Invalid email"}, 404

    return {"message": "login page"}


@main.route('/enter_code', methods=['GET', 'POST'])
def enter_code():
    if request.method == 'POST':
        data = request.get_json()
        code = data['code']

        if not code:
            return {"error": "Code is missing"}, 400

        _code = LoginData.query.filter_by(code=code).first()
        # Перевірка і логіка для перевірки коду
        if _code:
            # Перенаправлення на успішну сторінку авторизації
            return {"message": "code is valid. Go to profile page"}, 200
        else:
            return {"error": "Invalid code"}, 404

    return {"message": "enter code page"}


@main.route('/profile', methods=['GET', 'POST'])
def profile():
    auth_token = request.args.get('auth_token')
    client = Client.query.filter_by(auth_token=auth_token).first()

    if client:
        user = db.session.query(Client.first_name, Client.last_name, Client.phone, Client.email,
                                Client.client_code).filter_by(
            id=client.id).first()  # add photo. If exist get none

        response_data = {
            "message": "Auth is done",
            "user": get_formatted_user(user)
        }

        return jsonify(response_data), 200
   
    return jsonify({"error": "wrong auth token"}, 404)


import base64


@main.route('/upload_photo', methods=['GET', 'POST'])
def upload_photo():
    data = request.get_json()
    auth_token = data['auth_token']
    photo_base64 = data['photo']

    client = Client.query.filter_by(auth_token=auth_token).first()

    if client:
        # Convert the base64-encoded photo to bytes
        photo_bytes = base64.b64decode(photo_base64)

        # Save the photo to the database
        client.photo = photo_bytes
        db.session.commit()

        response_data = {
            "message": "Profile photo uploaded successfully",
            "photo": photo_bytes
        }

        return jsonify(response_data), 200
    else:
        return {"message": "photo not uploaded"}, 404


# @main.route("/logout", methods=['GET', 'POST'])
# def logout():#done
#     logout_user()
#     flash("Successfuly logged out.", "success")
#     return redirect(url_for('client.login'))


@main.route('/refresh', methods=['GET', 'POST'])
def refresh_token():
    auth_token = request.args.get('auth_token')
    client = Client.query.filter_by(auth_token=auth_token).first()

    if client is None:
        return {"message": "client not found"}, 404

    _t = verify_token(client.auth_token)

    if _t:
        updated_token = update_token_expiry(client.auth_token)
        if updated_token:
            client.auth_token = updated_token
            db.session.commit()

            response_data = {
                "message": "token is updated",
                "token": updated_token
            }
            return jsonify(response_data), 200
        else:
            return {"message": "token update failed"}, 500
    else:
        return {"message": "token is expired, login again"}, 401


import pandas


@main.route('/upload_clients', methods=['GET', 'POST'])
def upload_clients_from_excel():
    if request.method == 'GET':
        # Render the upload form
        return render_template('upload-excel.html')
    elif request.method == 'POST':
        # Read the File using Flask request
        file = request.files['file']
        # Save file in local directory
        file.save(file.filename)

        # Parse the data as a Pandas DataFrame type
        data = pandas.read_excel(file)

        # Iterate over the rows in the DataFrame and insert them into the database
        for row in data.itertuples(index=False):
            new_client = Client(
                first_name=row.first_name,
                last_name=row.last_name,
                phone=row.phone,
                email=row.email,
                photo=None,
                client_code=generate_code(),
                type_id=get_id(UserType, row.type),
                auth_token=''
            )

            db.session.add(new_client)

        # Commit the changes to the database
        db.session.commit()

        # Return HTML snippet that will render the table
        return data.to_html()


def get_formatted_user(user):
    return {'first_name': user.first_name, 'last_name': user.last_name,
                    'phone': user.phone, 'email': user.email, 'code': user.client_code, 'token': user.auth_token}#, 'code': user.client_code


def is_valid_email(email):
    # Регулярний вираз для перевірки формату електронної пошти
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'

    # Перевірка відповідності введеного email регулярному виразу
    if re.match(pattern, email):
        return True
    else:
        return False


def is_code_expired(due_time):
    current_time = datetime.utcnow()
    return current_time > due_time


@main.route('/loginn', methods=['GET', 'POST'])
def loginn():
    if request.method == 'POST':
        email = request.form.get('email')
        delete_expired_records()
        # Перевірка і логіка для перевірки email
        client = Client.query.filter_by(email=email).first()
        if client:
            # Видаляються записи яким більше 1 години
            delete_expired_records()

            existing_code = LoginData.query.filter_by(client_id=client.id).order_by(
                LoginData.creation_time.desc()).first()

            if existing_code and not is_code_expired(existing_code.due_time):
                # Якщо користувач вже має активний код, надсилаємо його замість генерації нового коду
                send_esputnik_email(email, existing_code.code)
            else:
                add_code(email)
                new_code = LoginData.query.order_by(LoginData.code.desc()).first()
                send_esputnik_email(email, new_code.code)

            # generate auth token
            auth_token = generate_token(client.id)
            # update client auth token in the database
            client.auth_token = auth_token
            db.session.commit()
            # login the user
            login_user(client)

            # Перенаправлення на сторінку введення коду
            return redirect(url_for('client.enter_codee'))
        else:
            return {"message": "wrong email"}

    return render_template('login.html')


@main.route('/enter_codee', methods=['GET', 'POST'])
def enter_codee():

    if request.method == 'POST':
        code = request.form.get('code')

        _code = LoginData.query.filter_by(code=code).first()
        # Перевірка і логіка для перевірки коду
        if _code:
            # Перенаправлення на успішну сторінку авторизації
            return redirect(url_for('client.successs'))
        else:
            return {"message": "wrong code"}

    return render_template('code.html')


@main.route('/successs')
def successs():
    user = db.session.query(Client.first_name, Client.last_name, Client.phone, Client.email, Client.client_code, Client.auth_token).filter_by(
        id=current_user.id).first()

    return get_formatted_user(user)


from datetime import datetime


def delete_expired_records():
    current_time = datetime.now()
    expired_records = LoginData.query.filter(LoginData.due_time <= current_time).all()
    for record in expired_records:
        db.session.delete(record)
    db.session.commit()


from esputnik.esputnik import ESputnikAPIAdaptor
import requests


def send_esputnik_email(email, code):
    url = 'https://esputnik.com/api/v1/message/email'
    # api_key = '26D8C43510E3919BB5371B087C5AB0FF'
    auth_key = "ZnVnYXM6MjZEOEM0MzUxMEUzOTE5QkI1MzcxQjA4N0M1QUIwRkY="

    headers = {
        'Authorization': f'Basic {auth_key}',
        'Content-Type': 'application/json'
    }

    payload = {
        "htmlText": f"<h4>Hello!</h4><br><p>Your login code is {code}</p>",
        "from": "itcluster.rv@gmail.com",
        "subject": "Your verification code for Rivne IT Card",
        "emails": [email]
    }

    try:
        response = requests.post(url, headers=headers, json=payload)
        response.raise_for_status()
        print('Email sent successfully')
    except requests.exceptions.HTTPError as errh:
        print(f'HTTP Error: {errh}')
    except requests.exceptions.ConnectionError as errc:
        print(f'Error Connecting: {errc}')
    except requests.exceptions.Timeout as errt:
        print(f'Timeout Error: {errt}')
    except requests.exceptions.RequestException as err:
        print(f'Error: {err}')

import openpyxl

@main.route('/get_faq', methods=['GET', 'POST'])
def faq():
    file_path = "./flask_api/utils/text_info/faq.xlsx"

    try:
        questions, answers = parse_excel_file(file_path)
        faq_data = [{'question': a, 'answer': q} for a, q in zip(questions, answers)]

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    response = make_response(jsonify(faq_data))
    response.headers['Content-Type'] = 'application/json; charset=utf-8'
    return response


def parse_excel_file(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    questions = []
    answers = []

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        cell_value = row[0].value

        if cell_value:
            if sheet.row_dimensions[row[0].row].index % 2 == 1:
                questions.append(cell_value)
            else:
                answers.append(cell_value)

    return questions, answers


from flask import send_file


@main.route('/privacy', methods=['GET', 'POST'])
def privacy():
    file_path = "utils/text_info/privacy.rtf"

    return send_file(file_path, as_attachment=True)

@main.route('/term_of_use', methods=['GET', 'POST'])
def term_of_use():
    file_path = "utils/text_info/terms_of_use.rtf"

    return send_file(file_path, as_attachment=True)

# Endpoint for deleting a record
@main.route("/delete/<id>", methods=["DELETE"])
def client_delete(id):
    client = Client.query.get(id)
    db.session.delete(client)
    db.session.commit()

    return {"message": "delete successfuly"}
